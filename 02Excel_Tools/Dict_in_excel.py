# -*- coding: utf-8 -*-

import glob
import os
import time
from openpyxl import Workbook

dic = r'PATH'
wb = Workbook()
sheet = wb.active

sheet.cell(1,1,"Path")
sheet.cell(1,2,"Name")
sheet.cell(1,3,"Type")
sheet.cell(1,4,"Change")

i = 1
for file in glob.iglob(dic+"/**/*.*",recursive = True):
    i+=1
    sheet.cell(i,1,file)
    sheet.cell(i,2,file.split("\\")[-1])
    sheet.cell(i,3,file.split(".")[-1])
    try: sheet.cell(i,4,time.ctime(os.stat(file).st_mtime))   
    except: sheet.cell(i,4,"Error")
    if i % 1000 == 0: print(i)
       
try: os.remove(r'PATH')
except: pass
wb.save(r'PATH')
wb.close()